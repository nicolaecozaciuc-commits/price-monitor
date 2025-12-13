import re
import logging
import time
import json
import os
from urllib.parse import quote_plus
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
from playwright.sync_api import sync_playwright

app = Flask(__name__, template_folder='templates')
CORS(app)

log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)
logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(message)s', datefmt='%H:%M:%S')
logger = logging.getLogger('PriceMonitor')

DEBUG_DIR = '/root/monitor/debug'
os.makedirs(DEBUG_DIR, exist_ok=True)

# ============ DIMENSION VALIDATION (V10.7) ============
def extract_dimensions(text):
    """Extract dimensions: 180x80, 180×80, 180 x 80"""
    if not text:
        return []
    pattern = r'(\d+)\s*[x×]\s*(\d+)'
    matches = re.findall(pattern, text, re.IGNORECASE)
    dimensions = []
    for match in matches:
        dim = f"{match[0]}x{match[1]}"
        if dim not in dimensions:
            dimensions.append(dim)
    return dimensions

def normalize_dimensions(dim_list):
    """Normalize for comparison: [180x80] → [80x180]"""
    normalized = []
    for dim in dim_list:
        parts = dim.split('x')
        if len(parts) == 2:
            try:
                a, b = int(parts[0]), int(parts[1])
                sorted_dim = f"{min(a,b)}x{max(a,b)}"
                if sorted_dim not in normalized:
                    normalized.append(sorted_dim)
            except:
                pass
    return sorted(normalized)

def validate_dimensions(sku_name, snippet_text, threshold=0.7):
    """Compare SKU dimensions vs snippet dimensions"""
    sku_dims = extract_dimensions(sku_name)
    snippet_dims = extract_dimensions(snippet_text)
    
    if not sku_dims:
        return {'valid': True, 'reason': 'No dims in SKU'}
    
    if not snippet_dims:
        return {'valid': False, 'reason': 'No dims in snippet'}
    
    sku_dims_norm = normalize_dimensions(sku_dims)
    snippet_dims_norm = normalize_dimensions(snippet_dims)
    
    matches = [d for d in sku_dims_norm if d in snippet_dims_norm]
    match_rate = len(matches) / len(sku_dims_norm) if sku_dims_norm else 0
    is_valid = match_rate >= threshold
    
    return {
        'valid': is_valid,
        'reason': f"Match {len(matches)}/{len(sku_dims_norm)}" if is_valid else f"Mismatch {len(matches)}/{len(sku_dims_norm)}"
    }

# ============ V12.6 - EXTRACTOR UNIVERSAL CU PRIORITATE MAXIMĂ ============
def extract_instock_price(text):
    """V12.7 - Prețul cu 'In stock/În stoc' - PRIORITATE MAXIMĂ"""
    match = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})\s*RON\s*[·•]\s*[ÎI]n stoc(?:k)?', text, re.IGNORECASE)
    if match:
        price_str = match.group(1).replace('.', '').replace(',', '.')
        try:
            val = float(price_str)
            if 50 < val < 500000:
                return val
        except:
            pass
    return None

# ============ SPECIAL EXTRACTORS (V12.2) ============
def extract_foglia_price(text):
    """Foglia: PREȚ RON · In stock"""
    match = re.search(r'([\d.,]+)\s*RON\s*[·●]\s*(?:●\s*)?[ÎI]n stoc', text, re.IGNORECASE)
    if match:
        price = clean_price(match.group(1))
        if price > 0:
            return price
    match = re.search(r'([\d.,]+)\s*RON[^·]*[ÎI]n stoc', text, re.IGNORECASE)
    if match:
        price = clean_price(match.group(1))
        if price > 0:
            return price
    return None

def extract_bagno_price(text):
    """Bagno.ro: max price (main product) - OLD VERSION"""
    prices = []
    matches = re.finditer(r'([\d.,]+)\s*(?:RON|Lei)', text, re.IGNORECASE)
    for match in matches:
        price = clean_price(match.group(1))
        if price > 0:
            prices.append(price)
    return max(prices) if prices else None

def extract_bagno_price_fixed(text):
    """V12.0 FIX: MIN price (first/main product) instead of MAX"""
    prices = []
    matches = re.finditer(r'([\d.,]+)\s*(?:RON|Lei)', text, re.IGNORECASE)
    for match in matches:
        price = clean_price(match.group(1))
        if price > 0:
            prices.append(price)
    return min(prices) if prices else None

def extract_germanquality_price(text):
    """Germanquality.ro: max price (main product) - OLD VERSION"""
    prices = []
    matches = re.finditer(r'([\d.,]+)\s*(?:RON|Lei)', text, re.IGNORECASE)
    for match in matches:
        price = clean_price(match.group(1))
        if price > 0:
            prices.append(price)
    return max(prices) if prices else None

def extract_germanquality_price_fixed(text):
    """V12.1 FIX: MIN price (main product) instead of MAX"""
    prices = []
    matches = re.finditer(r'([\d.,]+)\s*(?:RON|Lei)', text, re.IGNORECASE)
    for match in matches:
        price = clean_price(match.group(1))
        if price > 0:
            prices.append(price)
    return min(prices) if prices else None

def extract_neakaisa_price(text):
    """Neakaisa: max price (main product)"""
    prices = []
    matches = re.finditer(r'([\d.,]+)\s*(?:RON|Lei)', text, re.IGNORECASE)
    for match in matches:
        price = clean_price(match.group(1))
        if price > 0:
            prices.append(price)
    return max(prices) if prices else None

def extract_sensodays_price_fixed(text):
    """V12.2 FIX: MIN price for SensoDays"""
    prices = []
    matches = re.finditer(r'([\d.,]+)\s*(?:RON|Lei)', text, re.IGNORECASE)
    for match in matches:
        price = clean_price(match.group(1))
        if price > 0:
            prices.append(price)
    return min(prices) if prices else None

def filter_single_source_arhitecthuro(results):
    """V11.1 - Elimina arhitecthuro.ro daca apare DOAR intr-o singura sursa"""
    arhitecthuro_methods = [r['method'] for r in results if r.get('name') == 'arhitecthuro.ro']
    if len(arhitecthuro_methods) == 1:
        results = [r for r in results if r.get('name') != 'arhitecthuro.ro']
        logger.info(f"   🔻 Arhitecthuro filtered (single source)")
    return results

BLOCKED = ['u003e', 'google', 'bing', 'microsoft', 'facebook', 'youtube', 'doarbai', 'termohabitat', 'wikipedia', 'amazon', 'ebay', 'compari.ro']

SEARCH_URLS = {
    'emag.ro': 'https://www.emag.ro/search/{}',
    'absulo.ro': 'https://www.absulo.ro/catalogsearch/result/?q={}',
    'germanquality.ro': 'https://www.germanquality.ro/catalogsearch/result/?q={}',
    'sensodays.ro': 'https://www.sensodays.ro/catalogsearch/result/?q={}',
    'foglia.ro': 'https://www.foglia.ro/catalogsearch/result/?q={}',
    'bagno.ro': 'https://www.bagno.ro/c?query={}',
    'romstal.ro': 'https://www.romstal.ro/cautare?q={}',
    'compari.ro': 'https://www.compari.ro/search/?q={}',
    'ideal-standard.ro': 'https://www.ideal-standard.ro/ro/search?text={}',
    'instalatiiaz.ro': 'https://www.instalatiiaz.ro/?s={}',
    'dedeman.ro': 'https://www.dedeman.ro/ro/cautare?query={}',
    'baterii-lux.ro': 'https://www.baterii-lux.ro/cautare?controller=search&s={}',
    'badehaus.ro': 'https://www.badehaus.ro/cautare?search={}',
    'vasetoaleta.ro': 'https://www.vasetoaleta.ro/search?q={}',
    'decostores.ro': 'https://www.decostores.ro/search?q={}',
}

def clean_price(value):
    if not value: return 0
    text = re.sub(r'[^\d,.]', '', str(value))
    if not text: return 0
    if ',' in text and '.' in text:
        text = text.replace('.', '').replace(',', '.') if text.rindex(',') > text.rindex('.') else text.replace(',', '')
    elif ',' in text:
        text = text.replace(',', '.')
    try:
        price = float(text)
        return price if 50 < price < 500000 else 0
    except:
        return 0

def normalize(text):
    import unicodedata
    if not text: return ""
    text = unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]', '', text.lower())

def accept_cookies(page):
    selectors = [
        'button:has-text("Permite toate")',
        'button:has-text("Accept")',
        'button:has-text("Acceptă")',
        '#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll',
    ]
    for selector in selectors:
        try:
            btn = page.locator(selector).first
            if btn.is_visible(timeout=1000):
                btn.click(force=True)
                return True
        except:
            continue
    return False

def extract_prices_from_text(text):
    prices = []
    matches = re.findall(r'([\d.,]+)\s*Lei', text, re.IGNORECASE)
    for m in matches:
        p = clean_price(m)
        if p > 0 and p not in prices:
            prices.append(p)
    return prices[:10]

# ============ METODA 3: EXTRACȚIE HTML STRUCTURAT (RAFINATĂ) ============
def extract_from_google_html(page, sku):
    """Extrage prețuri din HTML - SKIPEAZĂ site-urile din BLOCKED"""
    results = []
    try:
        html_content = page.content()
        with open(f"{DEBUG_DIR}/google_{sku}_html.html", 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        price_patterns = re.finditer(
            r'([a-z0-9-]+\.ro)[^<>]{0,200}?([\d.,]+)\s*(?:RON|Lei)',
            html_content,
            re.IGNORECASE
        )
        
        for match in price_patterns:
            domain = match.group(1).lower()
            price = clean_price(match.group(2))
            
            if any(b in domain for b in BLOCKED):
                continue
            
            if not domain or len(domain) < 5:
                continue
            if price <= 0:
                continue
            if any(r['domain'] == domain for r in results):
                continue
            
            context = match.group(0).lower()
            transport_words = ['delivery', 'transport', 'livrare', 'shipping', 'expediere', ' sh']
            is_transport = any(tw in context for tw in transport_words)
            
            if not is_transport:
                results.append({'domain': domain, 'price': price, 'source': 'Google HTML'})
                logger.info(f"      🟠 {domain}: {price} Lei (HTML)")
        
        price_patterns_rev = re.finditer(
            r'([\d.,]+)\s*(?:RON|Lei)[^<>]{0,200}?([a-z0-9-]+\.ro)',
            html_content,
            re.IGNORECASE
        )
        
        for match in price_patterns_rev:
            price = clean_price(match.group(1))
            domain = match.group(2).lower()
            
            if any(b in domain for b in BLOCKED):
                continue
            
            if not domain or len(domain) < 5:
                continue
            if price <= 0:
                continue
            if any(r['domain'] == domain for r in results):
                continue
            
            context = match.group(0).lower()
            transport_words = ['delivery', 'transport', 'livrare', 'shipping', 'expediere', ' sh']
            is_transport = any(tw in context for tw in transport_words)
            
            if not is_transport:
                results.append({'domain': domain, 'price': price, 'source': 'Google HTML'})
                logger.info(f"      🟠 {domain}: {price} Lei (HTML)")
        
        if results:
            logger.info(f"   🟠 Metoda HTML: {len(results)} găsite")
    
    except Exception as e:
        logger.info(f"   ⚠️ HTML extract: {str(e)[:40]}")
    
    return results

def google_stealth_search(page, query, sku_for_match=None, sku_name=None, add_price_suffix=True):
    """Google search cu Metoda 1 (line), Metoda 2 (bloc), Metoda 3 (HTML)"""
    results = []
    search_query = f"{query} pret RON" if add_price_suffix else query
    url = f"https://www.google.com/search?q={quote_plus(search_query)}&hl=ro&gl=ro"
    file_suffix = sku_for_match or query.replace(' ', '_')[:20]
    
    try:
        page.goto(url, timeout=15000, wait_until='domcontentloaded')
        time.sleep(2)
        
        try:
            page.click('button:has-text("Accept all")', timeout=2000)
        except:
            try:
                page.click('button:has-text("Acceptă tot")', timeout=1000)
            except:
                pass
        
        time.sleep(1)
        page.screenshot(path=f"{DEBUG_DIR}/google_{file_suffix}.png")
        
        body_text = page.locator('body').inner_text()
        with open(f"{DEBUG_DIR}/google_{file_suffix}.txt", 'w', encoding='utf-8') as f:
            f.write(body_text)
        
        lines = body_text.split('\n')
        current_domain = None
        
        for i, line in enumerate(lines):
            line_lower = line.lower()
            
            domain_match = re.search(r'(?:https?://)?(?:www\.)?([a-z0-9-]+\.ro)', line_lower)
            if domain_match:
                d = domain_match.group(1)
                if len(d) > 4 and not any(b in d for b in BLOCKED):
                    current_domain = d
            
            query_lower = query.lower()
            has_match = query_lower in line_lower
            if not has_match and len(query.split()) > 1:
                words = query_lower.split()
                matches = sum(1 for w in words if w in line_lower and len(w) > 3)
                has_match = matches >= 2
            
            if has_match and current_domain:
                context = ' '.join(lines[max(0,i-2):min(len(lines),i+3)])
                
                # ============ V12.6 PRIORITATE MAXIMĂ: Preț cu "In stock" ============
                instock_price = extract_instock_price(context)
                if instock_price and instock_price > 0:
                    if not any(r['domain'] == current_domain for r in results):
                        results.append({'domain': current_domain, 'price': instock_price, 'source': 'Google SERP (InStock)'})
                        logger.info(f"      🟢 {current_domain}: {instock_price} Lei (InStock)")
                    continue
                
                # SPECIAL GERMANQUALITY
                if current_domain == 'germanquality.ro':
                    gq_price = extract_germanquality_price_fixed(context)
                    if gq_price and gq_price > 0:
                        if not any(r['domain'] == current_domain for r in results):
                            results.append({'domain': current_domain, 'price': gq_price, 'source': 'Google SERP (GQ)'})
                            logger.info(f"      🟠 {current_domain}: {gq_price} Lei (GQ)")
                        continue
                
                # SPECIAL FOGLIA
                if current_domain == 'foglia.ro':
                    foglia_price = extract_foglia_price(context)
                    if foglia_price and foglia_price > 0:
                        if not any(r['domain'] == current_domain for r in results):
                            results.append({'domain': current_domain, 'price': foglia_price, 'source': 'Google SERP (Foglia)'})
                            logger.info(f"      🟣 {current_domain}: {foglia_price} Lei (Foglia)")
                        continue
                
                # SPECIAL BAGNO
                if current_domain == 'bagno.ro':
                    bagno_price = extract_bagno_price_fixed(context)
                    if bagno_price and bagno_price > 0:
                        if not any(r['domain'] == current_domain for r in results):
                            results.append({'domain': current_domain, 'price': bagno_price, 'source': 'Google SERP (Bagno)'})
                            logger.info(f"      🟡 {current_domain}: {bagno_price} Lei (Bagno)")
                        continue
                
                # SPECIAL NEAKAISA
                if current_domain == 'neakaisa.ro':
                    neakaisa_price = extract_neakaisa_price(context)
                    if neakaisa_price and neakaisa_price > 0:
                        if not any(r['domain'] == current_domain for r in results):
                            results.append({'domain': current_domain, 'price': neakaisa_price, 'source': 'Google SERP (Neakaisa)'})
                            logger.info(f"      🟤 {current_domain}: {neakaisa_price} Lei (Neakaisa)")
                        continue
                
                # SPECIAL SENSODAYS
                if current_domain == 'sensodays.ro':
                    sensodays_price = extract_sensodays_price_fixed(context)
                    if sensodays_price and sensodays_price > 0:
                        if not any(r['domain'] == current_domain for r in results):
                            results.append({'domain': current_domain, 'price': sensodays_price, 'source': 'Google SERP (Sensodays)'})
                            logger.info(f"      🟢 {current_domain}: {sensodays_price} Lei (Sensodays)")
                        continue
                
                # Generic extraction
                price_patterns = re.finditer(r'([\d.,]+)\s*(?:RON|Lei|lei)', context, re.IGNORECASE)
                valid_prices = []
                for pm in price_patterns:
                    price_value = clean_price(pm.group(1))
                    if price_value <= 0:
                        continue
                    start = max(0, pm.start() - 25)
                    end = min(len(context), pm.end() + 15)
                    price_context = context[start:end].lower()
                    transport_words = ['delivery', 'transport', 'livrare', 'shipping', 'expediere']
                    is_transport = any(tw in price_context for tw in transport_words)
                    if not is_transport:
                        valid_prices.append(price_value)
                
                if valid_prices:
                    price = min(valid_prices)
                    if not any(r['domain'] == current_domain for r in results):
                        results.append({'domain': current_domain, 'price': price, 'source': 'Google SERP'})
                        logger.info(f"      🟢 {current_domain}: {price} Lei")
        
        logger.info(f"   📸 Google: {len(results)} cu preț")
        
        # ============ METODA 2: BLOC ============
        logger.info(f"   🔍 Metoda 2: bloc...")
        current_domain = None
        domain_line = -1
        
        for i, line in enumerate(lines):
            line_lower = line.lower()
            domain_match = re.search(r'(?:https?://)?(?:www\.)?([a-z0-9-]+\.ro)', line_lower)
            if domain_match:
                d = domain_match.group(1)
                if len(d) > 4 and not any(b in d for b in BLOCKED):
                    current_domain = d
                    domain_line = i
            
            # DIRECT GERMANQUALITY extraction
            if current_domain and current_domain == 'germanquality.ro' and domain_line >= 0 and i <= domain_line + 6:
                if not any(r['domain'] == current_domain for r in results):
                    block_start = domain_line
                    block_end = min(len(lines), domain_line + 7)
                    block_text = ' '.join(lines[block_start:block_end])
                    
                    # V12.6 PRIORITATE: In stock
                    instock_price = extract_instock_price(block_text)
                    if instock_price and instock_price > 0:
                        results.append({'domain': current_domain, 'price': instock_price, 'source': 'Google SERP (InStock)'})
                        logger.info(f"      🟢 {current_domain}: {instock_price} Lei (InStock)")
                        current_domain = None
                        domain_line = -1
                        continue
                    
                    gq_price = extract_germanquality_price_fixed(block_text)
                    if gq_price and gq_price > 0:
                        results.append({'domain': current_domain, 'price': gq_price, 'source': 'Google SERP (GQ)'})
                        logger.info(f"      🟠 {current_domain}: {gq_price} Lei (GQ)")
                        current_domain = None
                        domain_line = -1
                        continue
            
            if current_domain and domain_line >= 0 and i <= domain_line + 6:
                query_lower = query.lower()
                if query_lower in line_lower:
                    if any(r['domain'] == current_domain for r in results):
                        continue
                    
                    block_start = domain_line
                    block_end = min(len(lines), domain_line + 7)
                    block_text = ' '.join(lines[block_start:block_end])
                    
                    # V12.6 PRIORITATE: In stock
                    instock_price = extract_instock_price(block_text)
                    if instock_price and instock_price > 0:
                        results.append({'domain': current_domain, 'price': instock_price, 'source': 'Google SERP (InStock)'})
                        logger.info(f"      🟢 {current_domain}: {instock_price} Lei (InStock)")
                        current_domain = None
                        domain_line = -1
                        continue
                    
                    # FOGLIA
                    if current_domain == 'foglia.ro':
                        foglia_price = extract_foglia_price(block_text)
                        if foglia_price and foglia_price > 0:
                            results.append({'domain': current_domain, 'price': foglia_price, 'source': 'Google SERP (Foglia)'})
                            logger.info(f"      🟣 {current_domain}: {foglia_price} Lei (Foglia)")
                            current_domain = None
                            domain_line = -1
                            continue
                    
                    # BAGNO
                    if current_domain == 'bagno.ro':
                        bagno_price = extract_bagno_price_fixed(block_text)
                        if bagno_price and bagno_price > 0:
                            results.append({'domain': current_domain, 'price': bagno_price, 'source': 'Google SERP (Bagno)'})
                            logger.info(f"      🟡 {current_domain}: {bagno_price} Lei (Bagno)")
                            current_domain = None
                            domain_line = -1
                            continue
                    
                    # NEAKAISA
                    if current_domain == 'neakaisa.ro':
                        neakaisa_price = extract_neakaisa_price(block_text)
                        if neakaisa_price and neakaisa_price > 0:
                            results.append({'domain': current_domain, 'price': neakaisa_price, 'source': 'Google SERP (Neakaisa)'})
                            logger.info(f"      🟤 {current_domain}: {neakaisa_price} Lei (Neakaisa)")
                            current_domain = None
                            domain_line = -1
                            continue
                    
                    # SENSODAYS
                    if current_domain == 'sensodays.ro':
                        sensodays_price = extract_sensodays_price_fixed(block_text)
                        if sensodays_price and sensodays_price > 0:
                            results.append({'domain': current_domain, 'price': sensodays_price, 'source': 'Google SERP (Sensodays)'})
                            logger.info(f"      🟢 {current_domain}: {sensodays_price} Lei (Sensodays)")
                            current_domain = None
                            domain_line = -1
                            continue
                    
                    # Generic bloc extraction
                    price_patterns = re.finditer(r'([\d.,]+)\s*(?:RON|Lei|lei)', block_text, re.IGNORECASE)
                    valid_prices = []
                    for pm in price_patterns:
                        price_value = clean_price(pm.group(1))
                        if price_value <= 0:
                            continue
                        start = max(0, pm.start() - 25)
                        end = min(len(block_text), pm.end() + 15)
                        price_context = block_text[start:end].lower()
                        transport_words = ['delivery', 'transport', 'livrare', 'shipping', 'expediere']
                        is_transport = any(tw in price_context for tw in transport_words)
                        if not is_transport:
                            valid_prices.append(price_value)
                    
                    if valid_prices:
                        price = valid_prices[0]
                        results.append({'domain': current_domain, 'price': price, 'source': 'Google SERP (bloc)'})
                        logger.info(f"      🔵 {current_domain}: {price} Lei (bloc)")
                    
                    current_domain = None
                    domain_line = -1
        
        logger.info(f"   📸 Total după bloc: {len(results)}")
        
        # ========== METODA 3: HTML ==========
        html_results = extract_from_google_html(page, query)
        for r in html_results:
            if not any(existing['domain'] == r['domain'] for existing in results):
                results.append(r)
        
        if html_results:
            logger.info(f"   📸 Total după HTML: {len(results)}")
        
    except Exception as e:
        logger.info(f"   ⚠️ Google: {str(e)[:40]}")
    
    return results

def get_domains_from_bing(page, sku):
    """Bing fallback"""
    results = []
    try:
        for block in page.locator('.b_algo').all()[:15]:
            try:
                text = block.inner_text()
                text_lower = text.lower()
                
                domain = None
                for line in text.split('\n')[:3]:
                    match = re.search(r'(?:https?://)?(?:www\.)?([a-z0-9-]+\.ro)', line.lower())
                    if match:
                        d = match.group(1)
                        if len(d) > 4 and not any(b in d for b in BLOCKED):
                            domain = d
                            break
                
                if not domain:
                    continue
                if any(r['domain'] == domain for r in results):
                    continue
                
                has_sku = sku.lower() in text_lower
                price = 0
                price_match = re.search(r'([\d.,]+)\s*(?:RON|Lei|lei)', text)
                if price_match:
                    price = clean_price(price_match.group(1))
                
                results.append({'domain': domain, 'price': price, 'has_sku': has_sku, 'source': 'Bing SERP'})
                
                if price > 0 and has_sku:
                    logger.info(f"      🔵 {domain}: {price} Lei")
                elif has_sku:
                    logger.info(f"      🔵 {domain}: (pe site)")
                    
            except:
                continue
    except:
        pass
    
    return results

def find_price_on_site(page, domain, sku, save_debug=False):
    """Visit site if needed"""
    search_url = SEARCH_URLS.get(domain, f'https://www.{domain}/search?q={{}}')
    sku_norm = normalize(sku)
    sku_lower = sku.lower()
    url = search_url.format(quote_plus(sku))
    
    try:
        page.goto(url, timeout=15000, wait_until='domcontentloaded')
        time.sleep(3)
        
        if accept_cookies(page):
            time.sleep(2)
            page.reload(wait_until='domcontentloaded')
            time.sleep(3)
        
        page.evaluate("window.scrollTo(0, 500)")
        time.sleep(1)
        
        if save_debug:
            page.screenshot(path=f"{DEBUG_DIR}/{domain}_{sku}.png")
        
        body_text = page.locator('body').inner_text()
        body_lower = body_text.lower()
        
        error_phrases = ['0 produse', 'nu s-au gasit', 'nu am gasit', 'niciun rezultat', '0 rezultate']
        for phrase in error_phrases:
            if phrase in body_lower and 'produse)' not in body_lower:
                logger.info(f"         ⚠️ {phrase}")
                return None
        
        has_sku = sku_lower in body_lower or sku_norm in normalize(body_text)
        if not has_sku:
            return None
        
        prices = extract_prices_from_text(body_text)
        if prices:
            return {'price': prices[0], 'url': url}
        
        return None
        
    except Exception as e:
        logger.info(f"         ❌ {str(e)[:30]}")
        return None

def scan_product(sku, name, your_price=0):
    found = []
    sku = str(sku).strip()
    
    logger.info(f"🔎 {sku} - {name[:30]}...")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=['--disable-blink-features=AutomationControlled', '--no-sandbox']
        )
        
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={'width': 1920, 'height': 1080},
            locale='ro-RO',
            timezone_id='Europe/Bucharest',
        )
        
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        """)
        
        page = context.new_page()
        
        try:
            # ============ V13: Google #1 - SKU SIMPLU (fără "pret RON") - PRIMUL! ============
            logger.info(f"   🔍 Google #1: SKU simplu...")
            google_results_simple = google_stealth_search(page, sku, f"{sku}_simple", sku_name=name, add_price_suffix=False)
            
            for r in google_results_simple:
                if r['price'] > 0:
                    found.append({
                        'name': r['domain'],
                        'price': r['price'],
                        'url': f"https://www.{r['domain']}",
                        'method': 'Google Simple'
                    })
                    logger.info(f"      🔵 {r['domain']}: {r['price']} Lei (simplu)")
            
            # ============ Google #2: SKU + "pret RON" ============
            if len(found) < 5:
                logger.info(f"   🔍 Google #2: SKU + pret...")
                google_results = google_stealth_search(page, sku, sku, sku_name=name, add_price_suffix=True)
                
                for r in google_results:
                    if r['price'] > 0 and not any(f['name'] == r['domain'] for f in found):
                        found.append({
                            'name': r['domain'],
                            'price': r['price'],
                            'url': f"https://www.{r['domain']}",
                            'method': 'Google SKU'
                        })
                        logger.info(f"      🟢 {r['domain']}: {r['price']} Lei (SKU+pret)")
            
            # ============ Google #3: Denumire + "pret RON" ============
            if len(found) < 5 and name and len(name) > 10:
                logger.info(f"   🔍 Google #3: Denumire...")
                name_words = name.split()[:6]
                name_query = ' '.join(name_words)
                if sku.upper() not in name_query.upper():
                    name_query += f" {sku}"
                
                google_results_name = google_stealth_search(page, name_query, f"{sku}_name", sku_name=name)
                
                for r in google_results_name:
                    if r['price'] > 0 and not any(f['name'] == r['domain'] for f in found):
                        found.append({
                            'name': r['domain'],
                            'price': r['price'],
                            'url': f"https://www.{r['domain']}",
                            'method': 'Google Name'
                        })
                        logger.info(f"      🟡 {r['domain']}: {r['price']} Lei (din denumire)")
            
            if len(found) < 3:
                logger.info(f"   🔍 Bing completează...")
                query = f"{sku} pret"
                url = f"https://www.bing.com/search?q={quote_plus(query)}"
                
                page.goto(url, timeout=20000, wait_until='domcontentloaded')
                time.sleep(3)
                
                try:
                    page.click('#bnp_btn_accept', timeout=3000)
                except:
                    pass
                
                page.screenshot(path=f"{DEBUG_DIR}/bing_{sku}.png")
                
                bing_results = get_domains_from_bing(page, sku)
                
                for r in bing_results:
                    if any(f['name'] == r['domain'] for f in found):
                        continue
                    if r['price'] > 0 and r.get('has_sku'):
                        found.append({
                            'name': r['domain'],
                            'price': r['price'],
                            'url': f"https://www.{r['domain']}",
                            'method': 'Bing SERP'
                        })
            
            logger.info(f"   📊 Total: {len(found)}")
            
        except Exception as e:
            logger.info(f"   ❌ {str(e)[:50]}")
        finally:
            page.close()
        
        browser.close()
    
    for r in found:
        r['diff'] = round(((r['price'] - your_price) / your_price) * 100, 1) if your_price > 0 else 0
    
    if your_price > 0:
        before_filter = len(found)
        found = [r for r in found if -30 <= r['diff'] <= 30]
        filtered_count = before_filter - len(found)
        if filtered_count > 0:
            logger.info(f"   🔻 Filtrat {filtered_count} outliers (±30%)")
    
    found = filter_single_source_arhitecthuro(found)
    found.sort(key=lambda x: x['price'])
    return found[:5]

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/check', methods=['POST'])
def api_check():
    data = request.json
    your_price = float(data.get('price', 0) or 0)
    results = scan_product(data.get('sku', ''), data.get('name', ''), your_price)
    return jsonify({"status": "success", "competitors": results})

@app.route('/debug/<filename>')
def get_debug(filename):
    filepath = f"{DEBUG_DIR}/{filename}"
    if os.path.exists(filepath):
        return send_file(filepath)
    return "Not found", 404

# ============ EXCEL EXPORT (V12.6) ============
@app.route('/api/report', methods=['POST'])
def api_report():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from datetime import datetime
        
        data = request.json or {}
        products = data.get('products', [])
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raport Preturi"
        
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        sku_font = Font(bold=True, size=12)
        sku_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        
        ws.merge_cells('A1:E1')
        ws['A1'] = "RAPORT MONITORIZARE PRETURI"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'] = f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        row = 4
        for p in products:
            if not p.get('comps'): continue
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = f"SKU: {p.get('sku', '')} | {p.get('name', '')}"
            ws[f'A{row}'].font = sku_font
            ws[f'A{row}'].fill = sku_fill
            row += 1
            ws[f'A{row}'] = "Pret Nostru:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = f"{p.get('price', 0):.2f} Lei"
            ws[f'C{row}'] = f"Data: {datetime.now().strftime('%Y-%m-%d')}"
            row += 1
            for col, h in enumerate(['Competitor', 'Pret', 'Diferenta', 'Status', 'Metoda'], 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            for comp in p.get('comps', []):
                diff = comp.get('diff', 0)
                status = "▼ MAI IEFTIN" if diff < -5 else "▲ MAI SCUMP" if diff > 5 else "= EGAL"
                ws.cell(row=row, column=1, value=comp.get('name', ''))
                ws.cell(row=row, column=2, value=comp.get('price', 0))
                ws.cell(row=row, column=3, value=f"{'+' if diff > 0 else ''}{diff}%")
                ws.cell(row=row, column=4, value=status)
                ws.cell(row=row, column=5, value=comp.get('method', 'Google'))
                row += 1
            row += 1
        
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Raport_Preturi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    logger.info("🚀 PriceMonitor v13 - SKU simplu PRIMUL pe :8080")
    app.run(host='0.0.0.0', port=8080)
